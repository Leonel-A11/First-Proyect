import openpyxl 

def cliente_existe(cedula):
    wb = openpyxl.load_workbook("clientes.xlsx")
    ws = wb.active 
    rows = ws.iter_rows(min_row=2 , values_only=True)
    for row in rows:
        if row[0] == cedula:
            return True
    return False 

def producto_existe(producto):
    wb = openpyxl.load_workbook("inventario.xlsx")
    ws = wb.active
    rows = ws.iter_rows(min_row=2 , values_only=True)
    for row in rows:
        if row[0] == producto:
            return True
    return False

def registrarcliente():
    cedula = input("Ingrese la cedula del cliente: ")

    cedula2 = cliente_existe(cedula)
    if cedula2 == True:
        print ("\nEste cliente ya existe\n")
    else:
        nombre = input("Ingrese el nombre del cliente: ")
        wb = openpyxl.load_workbook("clientes.xlsx") 
        ws = wb.active
        ws.append([cedula, nombre])
        wb.save("clientes.xlsx") 
        print ("\nCliente registrado!\n")

def registrarproductos():
    codigo = input("Ingrese el codigo del producto: ")
    codigo2 = producto_existe(codigo)

    if codigo2 == True:
        print ("\nEste producto ya existe\n")
    else:
        nombre = input("Escriba el nombre del producto: ")
        precio = float(input("Escriba el precio del producto: "))
        stock = int(input("¿Cantidades disponible del producto?: "))
        wb = openpyxl.load_workbook("inventario.xlsx")
        ws = wb.active
        ws.append([codigo,nombre,precio,stock])
        wb.save("inventario.xlsx")
        print ("\nProducto Registrado!\n")

def ver_productos5():
    wb = openpyxl.load_workbook("inventario.xlsx")
    ws = wb.active
    rows = ws.iter_rows(min_row=2)
    lista= []
    for row in rows:
        nombre = row[1].value
        stock = row[3].value
        if stock <= 5 and stock > 0:
            lista.append((nombre,stock))
    
    for nombre,stock in lista:
        print(f"Producto: {nombre} - Stock: {stock}")

def mostrarinventario():
    wb = openpyxl.load_workbook("inventario.xlsx")
    ws = wb.active
    rows = ws.iter_rows(min_row=2 , values_only=True)
    for row in rows:
        print(f" {row[0]} - Producto: {row[1]} - Precio: {row[2]} - Stock: {row[3]}")
    wb.close()

def mostrarregistroventas():
    wb = openpyxl.load_workbook("ventas.xlsx")
    ws = wb.active
    rows = ws.iter_rows(min_row=2 , values_only=True)
    for row in rows:
        print(f" numero:{row[0]} - fecha: {row[1]} - cedula: {row[2]} - producto: {row[3]} - cantidad: {row[4]} - total:{row[5]}")
    wb.close()

def actualizarproducto():
    codigo = input("Ingrese el codigo del producto a actualizar")

    wb = openpyxl.load_workbook("inventario.xlsx")
    ws = wb.active
    rows = ws.iter_rows(min_row=2)
    productoencontrado= False

    for row in rows:
        if row[0].value == codigo:
            productoencontrado = True
            nombre = row[1].value
            stock = row[3].value
            print (f"Producto: {nombre} - Stock: {stock} unidades")

            agregar = int(input("Cuantas unidades desea añadir?: "))
            row[3].value = stock + agregar

            wb.save("inventario.xlsx")
            print ("Inventario actualizado")
            break

    if productoencontrado == False:
        print ("Producto no encontrado")

def productoeninventario(codigo):
    wb = openpyxl.load_workbook("inventario.xlsx")
    ws = wb.active
    rows = ws.iter_rows(min_row=2)
    for row in rows:
        if row[0].value == codigo:
            return wb, ws, row 
    return None,None,None

def guardarventa(numero,fecha,cedula,productos,totalc,totalp):
    wb = openpyxl.load_workbook("ventas.xlsx")
    ws = wb.active
    ws.append([numero,fecha,cedula,productos,totalc,totalp])
    wb.save("ventas.xlsx")
    print("\nVenta Registrada con exito\n")

def nrodeventa():
    wb = openpyxl.load_workbook("ventas.xlsx")
    ws = wb.active
    ultimafila = ws.max_row
    if ultimafila > 1:
        numero = ultimafila
    else:
        numero = 1
    
    return numero

def venta():
    cedula = input("Escriba la cedula del cliente: ")
    if cliente_existe(cedula) == False:
        print("\nRegiste primero al cliente\n")
    else:
        mostrarinventario()

        productos = ""
        totalcan = 0
        totalp = 0

        while True:
            codigo = input("\n\nIngrese el codigo del producto o escriba 'X' para salir: ")
            if codigo.lower() == "x":
                print("Gracias por su compra")
                break

            wb, ws, row = productoeninventario(codigo)
            if row == None:
                print("\nProducto no encontrado\n")
                continue

            nombre = row[1].value
            precio = row[2].value
            stock = row[3].value

            cantidad = int(input("\nCantidades que lleva: "))

            if cantidad > stock:
                print ("\nCantidades insuficientes\n")
                continue

            row[3].value = row[3].value - cantidad
            wb.save("inventario.xlsx")

            productos += f"{nombre}({cantidad}), "
            totalcan = totalcan + cantidad
            totalp = totalp + (cantidad * precio)

            if totalcan == 0:
                print ("\nNo se registro venta\n")
            else:
                fecha = input ("\ningrese la fecha: ")
                numero = nrodeventa()
                guardarventa(numero,fecha,cedula,productos,totalcan,totalp)

                print(f"\n\nVenta registrada con exito, Total: $ {totalp}\n\n ")


